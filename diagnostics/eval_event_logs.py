import pandas as pd
import numpy as np
import os
import sys
from tqdm import tqdm 
from pathlib import Path
import warnings
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import argparse

from scipy.stats import ks_2samp, wasserstein_distance
from scipy.cluster.hierarchy import dendrogram, linkage, fcluster
from sklearn.preprocessing import StandardScaler

#make the 'utils' module discoverable
parent_dir = Path(__file__).resolve().parent.parent
sys.path.append(str(parent_dir))

from utils.helper import read_json, write_json, get_inter_arrival_times_from_list_of_timestamps
from source.case_arrival_distance_for_timestamps import case_arrival_distribution_distance_custom

warnings.filterwarnings('ignore')
    
class Evaluation:
    """
        Evaluates the performance of simulated timestamp data using various metrics and models.

        Parameters
        ----------
        save_path : str, optional
            Directory to save evaluation results and visualizations (default is None).
        total_runs : int, optional
            Number of simulation runs to evaluate (default is 1).
        metric : str, optional
            Evaluation metric to use, e.g., 'CADD' or 'day' (default is 'CADD').
        method_types : str, optional
            Type of methods to evaluate ('raw' or 'prob', default is 'raw').
        methods : list, optional
            Specific methods to evaluate. If not specified, all methods for the given method_type will be used.

        Attributes
        ----------
        root_dir : str
            Root directory of the project.
        simulations_path : str
            Path to the simulations data directory.
        dataset_names : dict
            Dictionary to store dataset names by evaluation method.
        test_data : dict
            Dictionary to store test data by evaluation method.
        sim_data : dict
            Dictionary to store simulation data by evaluation method.
        data_drawn : bool
            Indicates if simulation data has been processed.

        Methods
        -------
        get_sim_data(subfolder_path, run)
            Retrieves simulation data for a specific run.
        draw_simulated_n_true_data(eval_interarrivals=False)
            Loads and processes simulation and test data for evaluation.
        evaluate_model_performance(results_dict)
            Computes and displays evaluation metrics for each dataset and method.
        visualize_datetime_lists()
            Visualizes date occurrences for test and simulation data.
        get_date_frequency_df(times)
            Creates a DataFrame with counts of occurrences for each date.
        highlight_min(s)
            Highlights the minimum value in a DataFrame row.
    """
    def __init__(self, root_dir, save_path = None, total_runs = 1, metric = 'CADD', method_types = 'raw', methods=None):
        self.root_dir = root_dir
        if self.root_dir not in sys.path:
            sys.path.append(self.root_dir)
        self.simulations_path = os.path.join(self.root_dir, 'event_log_simulations')
        self.save_path = save_path

        # Define available methods for each type
        self.available_methods = {
            'raw': ['mean', 'exponential', 'best_distribution', 'npp', 'kde'],
            'prob': ['mean_prob', 'exponential_prob', 'best_distribution_prob', 'prophet', 'kde_prob', 'npp_prob']
            # 'prob': ['mean_prob', 'exponential_prob', 'best_distribution_prob', 'prophet', 'kde_prob', 'lstm', 'chronos', 'xgboost', 'npp_prob']
        }

        # Initialize method lists based on user input or defaults
        self.methods = methods if methods is not None else self.available_methods[method_types]
        print(f'self.methods: {self.methods}')
        # Validate methods
        invalid_methods = [m for m in self.methods if m not in self.available_methods[method_types]]
        if invalid_methods:
            raise ValueError(f"Invalid methods for {method_types}: {invalid_methods}. Available methods are: {self.available_methods[method_types]}")

        # Initialize data storage for each method
        self.dataset_names = {method: [] for method in self.methods}
        self.test_data = {method: [] for method in self.methods}
        self.sim_data = {method: [] for method in self.methods}

        self.method_types = method_types
        self.data_drawn = False
        self.total_runs = total_runs
        self.metric = metric

    def get_sim_data(self, subfolder_path, run):
        # Construct the file paths for simulated.json and test.json
        simulated_path = os.path.join(subfolder_path, f'simulated_run_{run}.json')
        sim_data = read_json(simulated_path)

        return sim_data
        
    def draw_simulated_n_true_data(self, eval_interarrivals=False):
        results_dict = {}
        progress_bar = tqdm(total=len(self.methods), desc='Drawing results from simulated data..', position=0)

        # iterate over each selected method
        for method in self.methods:
            folder_path = os.path.join(self.simulations_path, method)
            for subfolder in os.listdir(folder_path):
                print(f'current dataset: {subfolder}')
                subfolder_path = os.path.join(folder_path, subfolder)

                # if subfolder not in self.dataset_names[method]: #adds folders that yield not score: fix
                #     self.dataset_names[method].append(subfolder)

                if os.path.isdir(subfolder_path):
                    if subfolder not in results_dict:
                        results_dict[subfolder] = {}
                    
                    test_path = os.path.join(subfolder_path, 'test.json')
                    if not os.path.isfile(test_path):
                        #if no test.json is found, skip this folder 
                        continue
                    test_data = read_json(test_path)
                    self.test_data[method].append(test_data)

                    test_data_series = pd.Series(test_data.copy())
                    test_data_dt = pd.to_datetime(test_data_series, format="%d.%m.%Y %H:%M:%S")
                    test_data_dt_list = test_data_dt.tolist()

                    distances = {}
                    for run in range(1, self.total_runs + 1):
                        sim_data = self.get_sim_data(subfolder_path, run)
                        self.sim_data[method].append(sim_data)
                                
                        # transform string to datetime
                        sim_data_series = pd.Series(sim_data.copy())
                        sim_data_dt = pd.to_datetime(sim_data_series, format="%d.%m.%Y %H:%M:%S")
                        sim_data_dt_list = sim_data_dt.tolist()

                        #evaluate interarrival times if eval_interarrivals is True
                        emd_iat = None
                        if eval_interarrivals:
                            # Ensure the data is sorted
                            test_data_dt_list_sorted = sorted(test_data_dt_list)
                            sim_data_dt_list_sorted = sorted(sim_data_dt_list)
                            
                            #we need to ensure the simulated data spans the same date-range as the test data 
                            
                            # Step 1: Extract unique dates from test_data_dt_list_sorted
                            sim_start_date = sim_data_dt_list_sorted[0].date()
                            sim_end_date = sim_data_dt_list_sorted[-1].date()
                            
                            test_start_date = test_data_dt_list_sorted[0].date()
                            test_end_date = test_data_dt_list_sorted[-1].date()
                            delta_days = (test_end_date - test_start_date).days
                            test_dates = [test_start_date + timedelta(days=i) for i in range(delta_days + 1)]
                            
                            original_sim_len = len(sim_data_dt_list_sorted)

                            # Step 2: Filter sim_data_dt_list_sorted
                            sim_data_dt_list_filtered = [
                                dt for dt in sim_data_dt_list_sorted if dt.date() in test_dates
                            ]
                            # Number of values removed
                            num_removed = original_sim_len - len(sim_data_dt_list_filtered)
                            # Print key information
                            print(f"Number of values removed: {num_removed}")
                            print(f"Original sim data length: {original_sim_len}")
                            # Update sim_data_dt_list_sorted with the filtered data
                            sim_data_dt_list_sorted = sim_data_dt_list_filtered
                            print(f"New sim data length: {len(sim_data_dt_list_sorted)}\n")
                            # Recalculate sim_start and sim_end after filtering
                            if len(sim_data_dt_list_sorted) == 0:
                                print('simulated data is empty due to missed test range.')
                            else:
                                sim_start = sim_data_dt_list_sorted[0]
                                sim_end = sim_data_dt_list_sorted[-1]
                                test_start = test_data_dt_list_sorted[0]
                                test_end = test_data_dt_list_sorted[-1]

                                # Step 3: Recalculate overreach after filtering
                                if sim_start < test_start:
                                    overreach_start = test_start - sim_start
                                else:
                                    overreach_start = None

                                if sim_end > test_end:
                                    overreach_end = sim_end - test_end
                                else:
                                    overreach_end = None
                                if overreach_start:
                                    print(f"Simulated data starts {overreach_start} earlier than test data.")
                                else:
                                    print("Simulated data does not start before test data.")
                                if overreach_end:
                                    print(f"Simulated data ends {overreach_end} later than test data.")
                                else:
                                    print("Simulated data does not end after test data.")

                                # Print the new first and last timestamps
                                print('\nAfter truncation:')
                                print(f'first test timestamp: {test_data_dt_list_sorted[0]}')
                                print(f'final test timestamp: {test_data_dt_list_sorted[-1]}\n')
                                print('---------------------')
                                print(f'first sim timestamp: {sim_data_dt_list_sorted[0]}')
                                print(f'final sim timestamp: {sim_data_dt_list_sorted[-1]}')

                            # Compute interarrival times
                            # test_interarrival_times = np.diff(test_data_dt_list_sorted)
                            # test_interarrival_times = [delta.total_seconds() for delta in test_interarrival_times]
                            # test_data_for_distance = test_interarrival_times
                            test_data_for_distance = get_inter_arrival_times_from_list_of_timestamps(test_data_dt_list_sorted)

                            
                            if len(sim_data_dt_list_sorted) == 0:
                                sim_interarrival_times = []
                                emd_iat = np.infty
                            else:
                                # sim_interarrival_times = np.diff(sim_data_dt_list_sorted)
                                # sim_interarrival_times = [delta.total_seconds() for delta in sim_interarrival_times]
                                # sim_data_for_distance = sim_interarrival_times
                                sim_data_for_distance = get_inter_arrival_times_from_list_of_timestamps(sim_data_dt_list_sorted)
                                if len(sim_data_for_distance) == 0:
                                    emd_iat = np.infty
                                else:
                                    emd_iat = wasserstein_distance(test_data_for_distance, sim_data_for_distance)
                        else:
                            # Use the arrival times directly
                            test_data_for_distance = test_data_dt_list
                            sim_data_for_distance = sim_data_dt_list

                        # evaluate
                        if eval_interarrivals:
                            print('Ignoring CADD and day_prob metric and calculate instead directly on interarrival-second level\n')
                            distances[f'run_{run}'] = np.sqrt(emd_iat)
                        else:
                            if self.metric == 'CADD':
                                bin = 'hour'
                            elif self.metric == 'day':
                                bin = 'day'
                            distance = case_arrival_distribution_distance_custom(
                                original_arrival_times=test_data_for_distance,
                                simulated_arrival_times=sim_data_for_distance,
                                bin=bin,
                                normalize=False,
                            )
                            distances[f'run_{run}'] = distance
                        print('calculations complete.\n')

                    # add result to dict
                    mean_dist = np.mean([d for d in distances.values()])
                    std_dist = np.std([d for d in distances.values()])

                    results_dict[subfolder][method] = {
                        'mean': np.round(mean_dist, 4),
                        'std': np.round(std_dist, 4)
                    }
                progress_bar.update(1)
        progress_bar.close()
        self.data_drawn = True

        return results_dict

    
    def evaluate_model_performance(self, results_dict):
        if not self.data_drawn:
            raise ValueError('Data needs to be drawn from simulations before evaluations can take place.')

        print('\nBegin evaluation of data..')
        rows = []
        mean_rows = []  # To hold only mean values for evaluation

        for dataset, values in results_dict.items():
            row_data = {'dataset': dataset}
            mean_row_data = {'dataset': dataset}

            # Add data for each method that was provided
            for method in self.methods:
                if method in values:  # Check if method exists in results
                    formatted_value = f"{values[method]['mean']} ({values[method]['std']})"
                    row_data[method] = formatted_value
                    mean_row_data[method] = values[method]['mean']

            rows.append(row_data)
            mean_rows.append(mean_row_data)

        df = pd.DataFrame(rows)
        mean_df = pd.DataFrame(mean_rows)  #this df is used for min calculations

        print('\nDisplaying top-performing simulation approaches..')
        print(df)
        
        # Find minimum values for each dataset
        for index, row in mean_df.iterrows():
            dataset_entry = row['dataset']
            # Get all columns except 'dataset' for finding minimum
            method_columns = [col for col in mean_df.columns if col != 'dataset']
            min_column = row[method_columns].idxmin()
            min_value = row[min_column]
            print(f"Dataset Entry: {dataset_entry}, Minimum Value: {min_value}, Respective Column: {min_column}")

        #save the styled DataFrame to an Excel file if save_path is provided
        if self.save_path is not None:
            file_path = self.save_path
            print(f'Saving simulation performance overview to {file_path}..')
            wb = Workbook()
            ws = wb.active

            for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(r)
            
            # Apply highlighting based on minimum values
            min_cols = mean_df.iloc[:, 1:].idxmin(axis=1)  # first col is 'dataset' which should not be considered
            header_row = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}  # Map header names to their column indices
            
            #highlight cells with minimum values
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df) + 1), 1):
                min_col_name = min_cols.iloc[idx-1]  # get the column name of the min value for the row
                min_col_index = header_row[min_col_name]  # get the Excel index
                row[min_col_index - 1].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")  # apply the fill

            # Create filename based on metric and method_types
            fn = f'performance_overview_simulations_{self.metric}_{self.method_types}.xlsx'
            wb.save(os.path.join(file_path, fn))
            print('Complete.\n')
    
#python eval_event_logs.py --res_dir=results
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Run evaluation of models on event log data.')
    parser.add_argument('--res_dir', type=str, default='results', help='Directory to save results to.')
    parser.add_argument('--total_runs', type=int, default=1, help='Amount of runs to be considered. Necessary for robust result construction.')
    parser.add_argument('--metric', type=str, default='CADD', help='Metric for evaluation.')
    parser.add_argument('--method_types', type=str, default='prob', help='If we evaluate the raw or prob methods.')
    parser.add_argument('--methods', nargs='+', help='Specific methods to evaluate. If not specified, all methods for the given method_type will be used.')
    args = parser.parse_args()

    save_path = os.path.join(os.getcwd(), args.res_dir)
    if not os.path.isdir(save_path):
        os.mkdir(save_path)
        
    eval_class = Evaluation(parent_dir, save_path, args.total_runs, args.metric, args.method_types, args.methods)
    
    results_dict = eval_class.draw_simulated_n_true_data()
    eval_class.evaluate_model_performance(results_dict)
    
    # if args.metric == 'CADD':
    # eval_class.visualize_datetime_lists()